import os
os.environ["CUDA_VISIBLE_DEVICES"] = '1'
from  transformers import AutoTokenizer
from  transformers import AutoModelForSeq2SeqLM
import torch
import math
import string
punctuation=string.punctuation
import nltk
from nltk.stem import PorterStemmer
ps = PorterStemmer()

import deepspeed
local_rank = int(os.getenv('LOCAL_RANK', '0'))
world_size = int(os.getenv('WORLD_SIZE', '1'))
device = torch.device("cuda", local_rank)
import numpy as np



trans_path="/home/yz/liukang/liukang/spanishpara/Helsinki-NLP/opus-mt-es-en"
model_name_or_path="/home/yz/liukang/liukang/spanishpara/Helsinki-NLP/nllb"

tokenizer = AutoTokenizer.from_pretrained(

    model_name_or_path, src_lang="eng_Latn"

)

model = AutoModelForSeq2SeqLM.from_pretrained(model_name_or_path, forced_bos_token_id=tokenizer.lang_code_to_id["eng_Latn"]).eval()
model=deepspeed.init_inference(model,
                                dtype=torch.float16,
                                mp_size=world_size,
                                replace_method="auto")

@torch.no_grad()
def trans_en_es(src_es):
    gen_kwargs_ori = {
        "max_length": 128,
        "num_beams": 5,
        "num_return_sequences":1,
        # "bad_words_ids":bad_words_ids
    }
    batch=trans_tokenizer(src_es,return_tensors="pt")
    input_ids=batch["input_ids"]
    attention_mask=batch["attention_mask"]

    tgt_en=trans_model.generate(
            batch["input_ids"],
            attention_mask=batch["attention_mask"],
            **gen_kwargs_ori
    )
    tgt_en=trans_tokenizer.decode(tgt_en[0],skip_special_tokens=True)
    return tgt_en
from pathlib import Path
import openpyxl
def getWordCount(word_count_path):
    word2count = {}
    xlsx_file = Path('',word_count_path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    last_column = sheet.max_column-1
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i==0:
            continue
        word2count[row[0]] = round(float(row[last_column]),3)
        
    return word2count

ori_sentences=open("/home/yz/liukang/liukang/fairseq-main_prefix/fairseq-main_prefix/data/tsar/tsar2022_en_test_gold.tsv",encoding="utf-8").readlines()
from nltk.tokenize.toktok import ToktokTokenizer
spanish_tok = ToktokTokenizer()
qutos=["</s>","<s>","<pad>"]
def skip_words(word):
    word=word.replace("eng_Latn","")
    word=word.replace("<s>","")
    word=word.replace("<pad>","")
    word=word.replace("</s>","")
    return word.strip()
from nltk.stem.snowball import SnowballStemmer
spanish_ps=SnowballStemmer("spanish")




words_frequency=getWordCount("/home/yz/liukang/liukang/fairseq-main_prefix/fairseq-main_prefix/data/tsar/SUBTLEX_frequency.xlsx")
import torch.nn.functional as F

from wordfreq import zipf_frequency





@torch.no_grad()
def give_real_scores_ahead(tokenizer,outputs,scores_with_suffix,scores_with_suffix_masks,suffix_tokens,prefix_len=None,prefix_str=None,max_ahead=1,flag=1,scores_no_suffix=None):
    beam_size,max_len=outputs.size()
    scores_with_suffix=scores_with_suffix[:,:max_len]
    scores_with_suffix_masks=scores_with_suffix_masks[:,:max_len]
    scores_no_suffix=scores_no_suffix[:,:max_len]

    first_index=prefix_len+2
    last_index=min(first_index+max_ahead,max_len)

    ahead_parts=outputs[:,1:]
    ahead_parts=ahead_parts.reshape(1,-1)[0].tolist()
    ahead_part_tokens=list(map(lambda x:tokenizer.convert_ids_to_tokens(x),ahead_parts))
    #ahead_part_tokens_masks=list(map(lambda x:not x.startswith("Ġ") and x not in qutos,ahead_part_tokens))
    ahead_part_tokens_masks=list(map(lambda x:not x.startswith("▁") and x not in qutos,ahead_part_tokens))
    ahead_part_tokens_masks=torch.tensor(ahead_part_tokens_masks)
    ahead_part_tokens_masks=ahead_part_tokens_masks.reshape(beam_size,-1)
    scores_with_suffix[:,:-1][ahead_part_tokens_masks]=-math.inf
    scores_with_suffix[scores_with_suffix_masks]=-math.inf 
    scores_no_suffix[:,:-1][ahead_part_tokens_masks]=-math.inf
    scores_no_suffix[scores_with_suffix_masks]=-math.inf 

    for j in range(0,first_index):
        scores_with_suffix[:,j]=torch.tensor(-math.inf)
        scores_no_suffix[:,j]=torch.tensor(-math.inf)

    for j in range(last_index,max_len):
        scores_with_suffix[:,j]=torch.tensor(-math.inf)   
        scores_no_suffix[:,j]=torch.tensor(-math.inf)  

    flat_scores_with_suffix=scores_with_suffix.reshape(1,-1).squeeze(dim=0)
    flat_scores_no_suffix=scores_no_suffix.reshape(1,-1).squeeze(dim=0)
    sorted_scores,sorted_indices=torch.topk(flat_scores_with_suffix,k=beam_size*max_ahead)


    beam_idx=sorted_indices//max_len
    len_idx=(sorted_indices%max_len)
    
    if flag!=None:
        #hope_len=len(spanish_tok.tokenize(prefix_str))+flag
        hope_len=len(nltk.word_tokenize(prefix_str))+flag
    else:
        hope_len=-1

    hope_outputs=[]
    hope_outputs_scores=[]
    no_suffix_scores=[]
    candis=[]

    for i in range(len(beam_idx)):
        if sorted_scores[i]==(-math.inf):
            continue

        #tmp_str1=tgt_dict.string(outputs[beam_idx[i],:(len_idx[i]+1)]).replace("@@ ","")
        #tmp_str1=tokenizer.decode(outputs[beam_idx[i],:(len_idx[i]+1)],skip_special_tokens=True)
        #tmp_str1="".join(tokenizer.convert_ids_to_tokens(outputs[beam_idx[i],:(len_idx[i]+1)])).replace("Ġ"," ")
        #tmp_str1=" ".join(tokenizer.convert_ids_to_tokens(outputs[beam_idx[i],:(len_idx[i]+1)])).replace(" ##","")
        tmp_str1=" ".join(tokenizer.convert_ids_to_tokens(outputs[beam_idx[i],:(len_idx[i]+1)])).replace(" ","").replace("▁"," ").strip()
        tmp_str1=skip_words(tmp_str1).strip()
        tmp_str1=tmp_str1.replace("<unk>","|")
        # if i==70:
        #     import pdb
        #     pdb.set_trace()
        #if len(spanish_tok.tokenize(tmp_str1))==hope_len:
        if len(nltk.word_tokenize(tmp_str1))==hope_len:
            # if tmp_str1.split()[-1]=="property":
            #     print(beam_idx[i])
            # print(tmp_str1.split()[-1])
            hope_outputs.append(outputs[beam_idx[i]])
            #print(tgt_dict.string(outputs[beam_idx[i]]),sorted_scores[i])
            hope_outputs_scores.append(sorted_scores[i].tolist())
            no_suffix_scores.append(scores_no_suffix[beam_idx[i],len_idx[i]].tolist())
            #import pdb;pdb.set_trace()
            #candis.append(spanish_tok.tokenize(tmp_str1)[-1].strip())

            tmp_candi=nltk.word_tokenize(tmp_str1)[len(nltk.word_tokenize(prefix_str)):]
            tmp_candi=" ".join(tmp_candi).strip()
            candis.append(tmp_candi)
            #candis.append(nltk.word_tokenize(tmp_str1)[-1].strip())

        elif hope_len==-1:
            # hope_outputs.append(outputs[beam_idx[i]])
            # hope_outputs_scores.append(sorted_scores[i].tolist())
            hope_outputs.append(outputs[beam_idx[i],:(len_idx[i]+1)])
            hope_outputs_scores.append(sorted_scores[i].tolist())
            tmp_candi=nltk.word_tokenize(tmp_str1)[len(nltk.word_tokenize(prefix_str)):]
            tmp_candi=" ".join(tmp_candi).strip()
            candis.append(tmp_candi)           
        # hope_outputs.append(outputs[beam_idx[i],:(len_idx[i]+1)])
        # hope_outputs_scores.append(sorted_scores[i].tolist())
        # if len(tmp_str1.split())==len(prefix_str.split())+1:
        #     print(tmp_str1)
    #print("&"*100)
    # import pdb
    # pdb.set_trace()
    return hope_outputs,hope_outputs_scores,candis,no_suffix_scores


@torch.no_grad()
def node_compute(model,tokenizer, sentence, complex_word,prefix,suffix,beam):
    sentence_tokens= tokenizer.encode(sentence, return_tensors='pt')
    prefix_tokens=tokenizer.encode(prefix, return_tensors='pt')[0][:-2]
    #prefix_tokens=tokenizer.encode(prefix)[:-2]
    #prefix_tokens=[25005,0]#spanish
    #suffix1=" ".join(spanish_tok.tokenize(suffix)[:3])
    suffix1=" ".join(nltk.word_tokenize(suffix)[:2])
    if suffix1.startswith(":"):
        suffix1=suffix1[1:].strip()
    #suffix1=""

    suffix_tokens=tokenizer.encode(suffix1.strip(), return_tensors='pt')[0][:-2].tolist()
    prefix_len=len(prefix_tokens)
    complex_tokens = tokenizer.encode(complex_word.strip())[:-2]
    attn_len = len(prefix_tokens)+len(complex_tokens)
    
    # prefix_tokens=None
    # suffix_tokens=None
    complex_first_tokens=tokenizer.encode(" ".join(complex_word.split()[:2]).strip())[:-2]
    outputs,scores_with_suffix,scores_with_suffix_masks,scores_no_suffix=model.generate(sentence_tokens.cuda(), 
    #outputs=model.generate(sentence_tokens.cuda(), 
                            num_beams=beam, 
                            min_length=3,
                            max_length=attn_len+2+20,
                            num_return_sequences=beam,
                            prefix_ids=prefix_tokens,
                            suffix_ids=suffix_tokens,
                            max_aheads=4,
                            tokenizer=tokenizer,
                            #complex_len=len(complex_first_tokens),
                            complex_len=1,
                            attn_len=None,
                            # return_dict_in_generate=True,
                            # output_scores=True
                        )
    if not torch.distributed.is_initialized() or torch.distributed.get_rank() == 0:
        outputs=outputs.cpu()
        scores_with_suffix=scores_with_suffix.cpu()
        scores_with_suffix_masks=scores_with_suffix_masks.cpu()
        scores_no_suffix=scores_no_suffix.cpu()
    #import pdb;pdb.set_trace()
    #print(scores_with_suffix)
    tmp_outputs,tmp_outputs_scores,tmp_candis,tmp_no_suffix_scores=give_real_scores_ahead(tokenizer,
                                                    outputs,
                                                    scores_with_suffix,
                                                    scores_with_suffix_masks,
                                                    suffix_tokens,
                                                    prefix_len=prefix_len,
                                                    prefix_str=prefix,
                                                    max_ahead=4,
                                                    flag=1,
                                                    scores_no_suffix=scores_no_suffix
                                                    )



    return tmp_candis, tmp_no_suffix_scores, tmp_outputs_scores

def node_compute_circle(model,tokenizer, sentence, complex_word,prefix,suffix,beam=-1):
    not_candis=["'ll",complex_word.lower(),"...",'']
    not_candis_words=['the','an','both',"was","he","had","have","his","a","which"]
    not_candis_words_first=[]
    not_candis_words_all=[]
    if(len(complex_word.split())==1):
        complex_lemma=ps.stem(complex_word.lower())
        not_candis.append(complex_lemma)
    if(len(suffix.strip())!=0):
        not_candis_words.append(suffix.strip().split()[0].strip())
    
    if(len(prefix.strip())!=0):
        not_candis_words_first.append(prefix.strip().split()[-1].strip())    

    prefix_list=[prefix]
    dict1={}
    prefix_len=len(nltk.word_tokenize(prefix))
    beam_list=[20,5,5]
    for step in range(3):
        all_prefix_list=[]
        beam=beam_list[step]
        for now_prefix in prefix_list:
            import time
            first_time=time.time()
            candis, candis_scores, candis_scores_with_suffix=node_compute(model,
                                                                        tokenizer, 
                                                                        sentence, 
                                                                        complex_word,
                                                                        now_prefix,
                                                                        suffix,
                                                                        beam)
            last_time=time.time()
            # print("use :",last_time-first_time)  
            for candi_i in range(len(candis)):
                candi_i_lemma=ps.stem(candis[candi_i].lower())
                if(candis[candi_i].lower() in punctuation or candis[candi_i].lower() in not_candis or candi_i_lemma in not_candis or candis[candi_i].isdigit()):
                    continue

                tmp_now_prefix=now_prefix+" "+candis[candi_i].lower()
                tmp_now_prefix=tmp_now_prefix.strip()
                if(tmp_now_prefix[len(prefix):].strip().lower() in dict1):
                    continue
                all_prefix_list.append(tmp_now_prefix)

                if(step!=0):

                    dict1[tmp_now_prefix[len(prefix):].strip().lower()]={
                            "scores_suffix":dict1[now_prefix[len(prefix):].strip().lower()]['scores_no_suffix']+candis_scores_with_suffix[candi_i],
                            "scores_no_suffix":dict1[now_prefix[len(prefix):].strip().lower()]['scores_no_suffix']+candis_scores[candi_i]
                    }
                else:
                    dict1[tmp_now_prefix[len(prefix):].strip().lower()]={
                            "scores_suffix":candis_scores_with_suffix[candi_i],
                            "scores_no_suffix":candis_scores[candi_i]
                    }    

        
        prefix_list=all_prefix_list
        #beam/=2
        # beam=int(beam)
    final_candis=[key1 for key1 in dict1 if key1.strip()!=complex_word]
    final_candis_scores=[dict1[key1]['scores_suffix'] for key1 in dict1 if key1.strip()!=complex_word]
    final_indices=torch.topk(torch.tensor(final_candis_scores),k=len(final_candis_scores),dim=0)[1]    
    final_indices=final_indices.tolist()
    
    sort_final_candis=[final_candis[index1] for index1 in final_indices]
    sort_final_candis_scores=[final_candis_scores[index1] for index1 in final_indices]
    # filter part
    part_words=complex_word.split()
    new_final_candis=[]
    new_final_candis_scores=[]
    for index1 in range(len(sort_final_candis)):
        if(sort_final_candis[index1] in part_words or sort_final_candis[index1].split()[-1] in not_candis_words or sort_final_candis[index1].split()[0].strip() in not_candis_words_first\
        or sort_final_candis[index1].endswith(complex_word)):
            continue
        if(len(suffix.strip().split())>1):
            first_suffix_two=" ".join(suffix.strip().split()[:2])
            first_suffix_two=first_suffix_two.strip()
            if(sort_final_candis[index1].strip().endswith(first_suffix_two)):
                continue
        new_final_candis.append(sort_final_candis[index1])
        new_final_candis_scores.append(sort_final_candis_scores[index1])


    #import pdb;pdb.set_trace()

    return new_final_candis[:100],new_final_candis_scores[:100]
            


@torch.no_grad()
def evaluate_lexical(model,tokenizer,file_reader,completed_steps):

    model.eval()
    potential=0
    import json
    f3=open("results/mwls.multibeam.json","w+")
    
    from tqdm import tqdm
    for i in tqdm(range(len(file_reader))):
        #original_text=file_reader[i][]
        target_word=file_reader[i]['target_word']

        prefix=file_reader[i]['prefix']
        prefix=" ".join(nltk.word_tokenize(prefix)).strip()
        again_prefix=" ".join(nltk.word_tokenize(prefix)).strip()
        if prefix!=again_prefix:
            print("prefix need again tokenize")
        prefix=again_prefix
        
        again_again_prefix=" ".join(nltk.word_tokenize(prefix))
        if again_again_prefix!=prefix:
            print("prefxi need again and again tokenize")
        prefix=again_again_prefix

        suffix=file_reader[i]['suffix']
        suffix=" ".join(nltk.word_tokenize(suffix)).strip()
        again_suffix=" ".join(nltk.word_tokenize(suffix)).strip()
        if suffix!=again_suffix:
            print("suffix need again tokenize")
        suffix=again_suffix

        again_again_suffix=" ".join(nltk.word_tokenize(suffix))
        if again_again_suffix!=suffix:
            print("suffix need again and again tokenize")
        suffix=again_again_suffix
        tmp_original_text=prefix+" "+target_word+" "+suffix
        tmp_original_text=tmp_original_text.strip()


        bert_substitutes, real_prev_scores=node_compute_circle(
            model,
            tokenizer,
            tmp_original_text,
            target_word,
            prefix,
            suffix,
            beam=8      
        )

        
        file_reader[i]["outputs"]=bert_substitutes
        file_reader[i]["outputs_scores"]=real_prev_scores
        write_str=json.dumps(file_reader[i])
        f3.write(write_str.strip()+"\n")
        f3.flush()

    f3.close()

    print("="*20)
    ndcg=0
    tp=0
    total_outputs=0
    total_labels=0
    potential=0
    for i1 in range(len(file_reader)):
        tp+=len(list(set(file_reader[i1]["outputs"][:5])&set(file_reader[i1]['labels'])))
        tmp_tp=len(list(set(file_reader[i1]["outputs"][:10])&set(file_reader[i1]['labels'])))
        total_outputs+=len(set(file_reader[i1]["outputs"][:5]))
        total_labels+=len(set(file_reader[i1]['labels']))

        if tmp_tp>0:
            potential+=1
        tmp_dcg=0
        for i2 in range(len(file_reader[i1]["outputs"])):
            if file_reader[i1]["outputs"][i2] in list(set(file_reader[i1]["outputs"])&set(file_reader[i1]['labels'])):
                tmp_dcg+=1/np.log2(i2+1+1)
        label_len=len(set(file_reader[i1]['labels']))
        label_dcg=0
        for i3 in range(label_len):
            label_dcg+=1/np.log2(i3+1+1)            
        ndcg+=tmp_dcg/label_dcg

    print("ndcg:",ndcg/len(file_reader))
    print("potential",potential/len(file_reader))
    print("precision",tp/total_outputs)
    print("recall",tp/total_labels)            

    return potential/len(ori_sentences)


file_reader=[]
file_lines=open("/home/yz/liukang/liukang/spanishpara/TASLP_results/data/MWLS1.tsv").readlines()

for line in file_lines:
    if line.strip().split('\t')[0]=="Id":
        continue
    tmp_dict={}
    #import pdb;pdb.set_trace()
    #import pdb;pdb.set_trace()
    try:
        tmp_dict["prefix"]=line.strip().split('\t')[2].lower()
        if(tmp_dict["prefix"].strip().endswith(",")):
            tmp_dict['prefix']=tmp_dict['prefix'][:-1].strip()

        tmp_dict["target_word"]=line.strip().split('\t')[3].lower()
        tmp_dict["suffix"]=line.strip().split('\t')[4].lower()
        tmp_labels=line.strip().split("\t")[6:]  
        if(len(tmp_labels)!=5):
            continue
        tmp_dict["labels"]={}
        for label1 in tmp_labels:
            if label1 not in tmp_dict["labels"]:
                tmp_dict["labels"][label1.lower()]=1
            else:
                tmp_dict["labels"][label1.lower()]+=1

        file_reader.append(tmp_dict)
    except:
        import pdb;pdb.set_trace()
print("total number",len(file_reader))
evaluate_lexical(model,tokenizer,file_reader,1)
